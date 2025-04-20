from django.shortcuts import render, redirect  # Renders HTML templates and redirects to a new view or URL
from django.contrib.auth.forms import UserCreationForm  # Form for creating a new user account
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth import login, logout, authenticate  # Functions for handling login, logout, and authentication
from django.contrib import messages  # Django messages framework for displaying feedback messages to users
from django.contrib.auth.decorators import login_required, user_passes_test  # Decorators for restricting access
from django.urls import reverse  # Generates URLs based on view names
from django.views.decorators.csrf import csrf_exempt  # Exempts a view from CSRF protection
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse  # HTTP responses
from django.views.decorators.cache import never_cache  # Prevents caching of the decorated view's response
from django.contrib.auth.hashers import make_password  # Hashes passwords securely
from django.core.mail import EmailMultiAlternatives, BadHeaderError, send_mail  # Email handling
from django.template.loader import render_to_string  # Renders templates to a string, useful for email templates
from django.views.decorators.http import require_POST, require_http_methods  # Restricts view methods
from django.core.files.base import ContentFile  # For working with file-like objects in memory
from django.conf import settings  # Accesses Django project settings

import logging  # Logging module for tracking events or errors
import json  # JSON data handling
import random  # Random number generation
import string  # Includes constants and utilities for string manipulation
import traceback  # For extracting and formatting exception tracebacks
import os  # OS utilities for interacting with the file system
import zipfile  # For handling zip archives
from io import BytesIO  # In-memory byte stream

import openpyxl  # Excel file handling
from openpyxl.styles import Font, Alignment, Border, Side  # Styling for Excel cells

import qrcode  # For generating QR codes
import pyotp  # For generating and verifying time-based OTPs

from .forms import RegistrationForm  # Custom registration form specific to your app
from .models import CustomUser, UserProfile, UploadedFileData, Entry  # Custom models
from .projects import get_project_entries, indexPage  # Custom project utility functions


def get_uploaded_files(request):
    project_code = request.GET.get('project_code')
    print({project_code})
    
    if project_code:
        files = UploadedFileData.objects.filter(project_code=project_code)
        print(f"Found {len(files)} files for project code {project_code}")

        file_data = [{
             'file_id': file.file_id,  
            'file_name': file.file_name,
            'file_type': file.file_type,
            'uploaded_at': file.uploaded_at.strftime('%Y-%m-%d %H:%M:%S'),
            'file_url': request.build_absolute_uri(file.file.url)  # 👈 Add the download URL here
        } for file in files]

        print(f"Files: {file_data}")
        
        return JsonResponse({'files': file_data})
    else:
        return JsonResponse({'files': []})

logger = logging.getLogger(__name__)  # Initializes logger for tracking events or errors
print("KwentasApp.views module loaded")  # Debugging print

def update_project(request, code):
    entry = Entry.objects.get(code=code)  # or whatever model you are using
    return render(request, 'procurements.html', {'entry': entry})


FILE_TYPE_KEYWORDS = {
    "PPMP": ["PPMP", "PROJECT PROCUREMENT MANAGEMENT PLAN"],
    "APP": ["APP", "ANNUAL PROCUREMENT PLAN"],
    "POW": ["POW", "PROGRAM OF WORKS"]  
}
# views.py

@csrf_exempt
def upload_excel(request):
    if request.method != 'POST':
        messages.error(request, "Invalid request method.")
        return redirect('procurements')

    uploaded_file = request.FILES.get('excel_file')
    selected_type = (request.POST.get('file_type') or '').upper()
    project_code  = request.POST.get('project_code')

    # Basic input checks
    if not uploaded_file:
        messages.error(request, "No file uploaded.")
        return redirect('procurements')

    if selected_type not in FILE_TYPE_KEYWORDS:
        messages.error(request, "Please select a valid file type (PPMP, APP, or POW).")
        return redirect('procurements')

    if not uploaded_file.name.lower().endswith(('.xlsx', '.xls')):
        messages.error(request, "Only Excel files are allowed.")
        return redirect('procurements')

    try:
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        keywords = FILE_TYPE_KEYWORDS[selected_type]

        found = False
        found_keywords = set()

        # search each sheet, row, cell
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if not cell:
                        continue
                    text = str(cell).upper()
                    for kw in keywords:
                        if kw in text:
                            found = True
                            found_keywords.add(kw)
                    if found:
                        break
                if found:
                    break
            if found:
                break

        if found:
            # Save record
            UploadedFileData.objects.create(
                project_code = project_code,
                file_type    = selected_type,
                file         = uploaded_file,
                file_name    = uploaded_file.name,
            )
            msgs = ", ".join(found_keywords)
            messages.success(
                request,
                f"Found keyword(s) [{msgs}] for {selected_type} — saved under project {project_code}."
            )
        else:
            messages.error(
                request,
                f"No '{selected_type}' or its full form found in the Excel file."
            )

    except Exception as e:
        messages.error(request, f"Error reading Excel file: {e}")

    return redirect('procurements')
def bulk_download_xlsx(request):
    if request.method == 'POST':
        selected_codes = request.POST.getlist('selected_entries')

        if not selected_codes:
            return JsonResponse({"error": "No entries selected"}, status=400)

        # Define the path to your pre-designed template
        template_path = os.path.join(settings.BASE_DIR, 'KwentasApp/static/KwentasApp/xls_templates/template_report.xlsx')

        # Load the pre-designed template
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # Define service types and initial row positions
        service_type_rows = {
            'General': 12,
            'Social': 12,
            'Economic': 19,
            'Environmental': 19,
        }
        max_entries = {'General': 5, 'Social': 5, 'Economic': 2, 'Environmental': 2}

        last_used_row = service_type_rows.copy()

        # Format row function excluding 'J' column
        def format_row(row_num):
            for cell in ws[row_num]:
                if cell.column_letter != 'J':  # Exclude the 'J' column
                    cell.font = Font(name='Calibri', size=10)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = Border(
                        left=Side(border_style="thin"),
                        right=Side(border_style="thin"),
                        top=Side(border_style="thin"),
                        bottom=Side(border_style="thin")
                    )

        # Initialize entry counters for each service type
        entry_counts = {key: 0 for key in service_type_rows.keys()}

        for code in selected_codes:
            _, _, all_entries = get_project_entries()
            selected_entry = next((entry for entry in all_entries if entry['code'] == code), None)

            if selected_entry:
                service_type = selected_entry.get('services', 'General')

                # Get the last used row for the service type
                start_row = last_used_row[service_type]

                # Write data directly into the existing row without inserting a new one
                ws.insert_rows(start_row)

                # Define cell positions based on the service type
                ppa_cell, location_cell, start_date_cell, end_date_cell, overall_budget_cell, total_disbursements_cell, remarks_cell = 'A', 'B', 'C', 'D', 'E', 'G', 'I'

                data = [
                    selected_entry.get('ppa', ''),
                    selected_entry.get('location', ''),
                    selected_entry.get('start_date', ''),
                    selected_entry.get('end_date', ''),
                    selected_entry.get('overall_budget', ''),
                    selected_entry.get('total_disbursements', ''),
                    selected_entry.get('remarks', ''),
                ]

                columns = [ppa_cell, location_cell, start_date_cell, end_date_cell, overall_budget_cell, total_disbursements_cell, remarks_cell]

                # Write data to the appropriate cells in the template
                for col_letter, value in zip(columns, data):
                    ws[f"{col_letter}{start_row}"].value = value

                # Format the newly added row
                format_row(start_row)

                # Update the last used row for the service type, ensuring no empty rows are skipped
                last_used_row[service_type] += 1
                entry_counts[service_type] += 1

        # Insert a new row after the last entry, regardless of service type
        ws.insert_rows(max(last_used_row.values()) + 1)

        # Set up the response as an Excel file download
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Projects Report.xlsx'

        wb.save(response)
        return response

    return JsonResponse({"error": "Invalid request"}, status=400)

#this def ensures the Sweet Alert in Homepage just shows after loggin in only
def unset_just_logged_in(request):
    if 'just_logged_in' in request.session:
        del request.session['just_logged_in']
    return JsonResponse({'status': 'success'})

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        remember_me = request.POST.get('remember_me') == 'on'
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            logger.info(f'Successful login for user: {username}')
            login(request, user)
            request.session['just_logged_in'] = True
            if remember_me:
                request.session.set_expiry(1209600)  # 2 weeks
            else:
                request.session.set_expiry(0)  # Expire on browser close
            try:
                # Check if the user has 2FA enabled
                user_profile = UserProfile.objects.get(user=user)
                if user_profile.two_factor_enabled:
                    request.session['otp_required'] = True  # Set flag in session
                    return redirect('verify_otp')
            except UserProfile.DoesNotExist:
                logger.warning(f'UserProfile does not exist for user: {username}')
                pass

            request.session['otp_verified'] = True  # Set OTP verified if no 2FA
            return redirect('homepage')

        else:
            logger.warning(f'Invalid login attempt for user: {username}')
            messages.error(request, 'Invalid credentials. Please try again.')
            return redirect('login')  # Redirect after a failed login attempt

    return render(request, 'KwentasApp/login.html')



def admin_view(request):
    return render(request, 'KwentasApp/admin')

@login_required
def base_view(request):
    logout(request)
    return redirect(reverse('login'))

def is_superuser(user):
    return user.is_authenticated and user.is_superuser

@csrf_exempt
def send_reset_code(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email')


            if not CustomUser.objects.filter(email=email).exists():
                return JsonResponse({'success': False, 'error': 'Email not found!'}, status=404)

            # ✅ Generate a 6-digit reset code
            code = ''.join(random.choices('0123456789', k=6))
            subject = 'Your Verification Code'
            text_content = f'Your verification code is: {code}'

            try:
                html_content = render_to_string('KwentasApp/verification_email.html', {'code': code})
            except Exception as e:
                logger.error(f'Error rendering email template: {str(e)}')
                return JsonResponse({'success': False, 'error': 'Error rendering email template.'}, status=500)

            try:
                msg = EmailMultiAlternatives(subject, text_content, 'kwentasklarasboljoon@gmail.com', [email])
                msg.attach_alternative(html_content, "text/html")
                msg.send()

                # ✅ Store the verification code and email in session
                request.session['verification_code'] = code
                request.session['email'] = email
                logger.info(f'Sent verification code to {email}')
                return JsonResponse({'success': True})
            except Exception as e:
                logger.error(f'Error sending email to {email}: {str(e)}')
                return JsonResponse({'success': False, 'error': 'Error sending email.'}, status=500)

        except json.JSONDecodeError:
            logger.warning('Invalid JSON format received in send_verification_code.')
            return JsonResponse({'success': False, 'error': 'Invalid JSON format.'}, status=400)

    logger.warning('Invalid request method for sending verification code.')
    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=405)

@csrf_exempt
def send_verification_code(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email')

            # ✅ Check if the email is already registered
            if CustomUser.objects.filter(email=email).exists():
                return JsonResponse({'success': False, 'error': 'This email is already registered!'}, status=400)

            # ✅ Generate a 6-digit numeric verification code
            code = ''.join(random.choices('0123456789', k=6))
            subject = 'Your Verification Code'
            text_content = f'Your verification code is: {code}'

            try:
                html_content = render_to_string('KwentasApp/verification_email.html', {'code': code})
            except Exception as e:
                logger.error(f'Error rendering email template: {str(e)}')
                return JsonResponse({'success': False, 'error': 'Error rendering email template.'}, status=500)

            try:
                msg = EmailMultiAlternatives(subject, text_content, 'kwentasklarasboljoon@gmail.com', [email])
                msg.attach_alternative(html_content, "text/html")
                msg.send()

                # ✅ Store the verification code and email in session
                request.session['verification_code'] = code
                request.session['email'] = email
                logger.info(f'Sent verification code to {email}')
                return JsonResponse({'success': True})
            except Exception as e:
                logger.error(f'Error sending email to {email}: {str(e)}')
                return JsonResponse({'success': False, 'error': 'Error sending email.'}, status=500)

        except json.JSONDecodeError:
            logger.warning('Invalid JSON format received in send_verification_code.')
            return JsonResponse({'success': False, 'error': 'Invalid JSON format.'}, status=400)

    logger.warning('Invalid request method for sending verification code.')
    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=405)

@user_passes_test(lambda u: u.is_superuser, login_url='login')
@csrf_exempt
def registration_view(request):
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            entered_code = request.POST.get('verification_code')
            # Retrieve the verification code and email from session
            generated_code = request.session.get('verification_code')
            session_email = request.session.get('email')
            entered_email = form.cleaned_data.get('email')

            # Check if the entered email matches the session email
            if entered_email != session_email:
                response = {
                    'status': 'error',
                    'message': 'Email mismatch. Please use the verified email.'
                }
                return JsonResponse(response)

            # Check if the verification code matches
            if entered_code == generated_code:
                try:
                    # Create the account if both email and code are correct
                    form.save()
                    # Clear session after successful registration
                    request.session.pop('verification_code', None)
                    request.session.pop('email', None)
                    response = {
                        'status': 'success',
                        'message': 'Your account has been created successfully!',
                    }
                    return JsonResponse(response)
                except Exception as e:
                    response = {
                        'status': 'error',
                        'message': f'An error occurred: {e}'
                    }
                    return JsonResponse(response)
            else:
                response = {
                    'status': 'error',
                    'message': 'Verification code does not match.',
                }
                return JsonResponse(response)
        else:
            print(form.errors)  # Log form errors for debugging
            response = {
                'status': 'error',
                'message': 'Form is invalid. Please correct the errors.',
                'errors': form.errors
            }
            return JsonResponse(response)
    else:
        form = RegistrationForm()
    return render(request, 'KwentasApp/register.html', {'form': form})

@user_passes_test(is_superuser, login_url='login')
def register_page(request):
    return render(request,'KwentasApp/register.html')

def logout_view(request):
    logout(request)
    return redirect(reverse('login'))

@login_required
def homepage(request):
    # Ensure 'just_logged_in' exists in the session
    if 'just_logged_in' not in request.session:
        request.session['just_logged_in'] = False

    # Check OTP requirement
    if request.session.get('otp_required') and not request.session.get('otp_verified'):
        messages.error(request, 'You must verify the OTP to access the homepage.')
        return redirect('verify_otp')

    user_name = request.user.username if request.user.is_authenticated else "Guest"


    print(f"Debug: user_name = {user_name}")  # Debugging output

    context = {
        'user_name': user_name,
        'just_logged_in': request.session.get('just_logged_in', False),  # Ensure it's passed
    }

    # Reset the flag after displaying the message
    request.session['just_logged_in'] = False

    return render(request, 'KwentasApp/homepage.html', context)
def forgotpassword(request):
    if request.method == 'POST':
        email = request.POST.get('email')

        # Check if email exists in the database
        try:
            user = CustomUser.objects.get(email=email)
        except ObjectDoesNotExist:
            # Added: Stop execution if the email is not registered
            messages.error(request, 'Email not found.')
            return redirect('forgotpassword') 
        # Generate a 6-digit OTP
        code = ''.join(random.choices('0123456789', k=6))

        # Save OTP and email in session
        request.session['verification_code'] = code
        request.session['reset_email'] = email

        # Send verification code to the email
        send_mail(
            'Password Reset Code',
            f'Your verification code is: {code}',
            'kwentasklarasboljoon@gmail.com',
            [email],
            fail_silently=False,
        )

        messages.success(request, 'Verification code sent to your email.')
        return redirect('verify_password_reset')

    return render(request, 'KwentasApp/forgot-password.html')
    
@csrf_exempt
def verify_and_change_password(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email')
            code = data.get('code')
            password = data.get('password')
            stored_code = request.session.get('verification_code')

            logger.info(f'Received request to change password for {email}')
            logger.info(f'Stored code: {stored_code}, Session email: {request.session.get("reset_email")}')

            # Ensure email exists before proceeding
            try:
                user = CustomUser.objects.get(email=email)  # Added: Ensure the email exists in the database
            except CustomUser.DoesNotExist:
                logger.error(f'User not found for email: {email}')
                return JsonResponse({'success': False, 'error': 'User not found'}, status=404)  # Added: Return error if user is not found

            # Check if OTP matches and email is correct
            if code and code == stored_code and email == request.session.get('reset_email'):
                hashed_password = make_password(password)
                user.password = hashed_password
                user.save()

                # Log successful password change
                logger.info(f'Password changed for user: {email}')

                # Clear session data
                del request.session['verification_code']
                del request.session['reset_email']

                # Attempt to log in the user with the new password
                user = authenticate(username=user.username, password=password)
                if user is not None:
                    login(request, user)
                    logger.info(f'User {email} logged in successfully after password change.')
                    return JsonResponse({'success': True})
                else:
                    logger.error(f'Authentication failed for user after password change: {email}')
                    return JsonResponse({'success': False, 'error': 'Authentication failed after password change'}, status=400)

            logger.warning(f'Invalid code or email mismatch for user: {email}')
            return JsonResponse({'success': False, 'error': 'Invalid code or email mismatch'}, status=400)

        except json.JSONDecodeError:
            logger.warning('JSON decode error during password change process.')
            return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
        except Exception as e:
            logger.error(f'Unexpected error during password change process: {str(e)}\n{traceback.format_exc()}')
            return JsonResponse({'success': False, 'error': 'Internal server error.'}, status=500)

    logger.warning('Invalid request method for password change.')
    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=400)

@login_required
def generate_qr_code(request):
    if request.method == 'POST':
        user_profile = request.user.userprofile  # Ensure userprofile exists
        
        if request.POST.get('confirm_2fa') == 'true':
            user_profile.two_factor_enabled = True
            
            # Generate new TOTP secret key
            secret_key = pyotp.random_base32()
            user_profile.totp_secret = secret_key
            
            # Create TOTP URL for authenticator apps
            totp = pyotp.TOTP(secret_key)
            otp_uri = totp.provisioning_uri(name=request.user.username, issuer_name='Kwentas Klaras')
            
            # Generate QR code
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4
            )
            qr.add_data(otp_uri)
            qr.make(fit=True)

            # Save QR code image
            img = qr.make_image(fill='black', back_color='white')
            buffer = BytesIO()
            img.save(buffer, format='PNG')
            user_profile.qr_code.save('qr_code.png', ContentFile(buffer.getvalue()))
            
            user_profile.save()
            
            return JsonResponse({'success': True, 'qr_code_url': user_profile.qr_code.url})
        else:
            user_profile.two_factor_enabled = False
            user_profile.save()
            return JsonResponse({'success': True})
    return JsonResponse({'success': False})

@login_required
def enable_2fa(request):
    if request.method == 'POST':
        user_profile = request.user.userprofile

        if request.POST.get('confirm_2fa') == 'true':
            user_profile.two_factor_enabled = True
        else:
            user_profile.two_factor_enabled = False
        
        if user_profile.two_factor_enabled:
            # Generate TOTP secret and URI
            totp = pyotp.TOTP(pyotp.random_base32())
            secret = totp.secret
            otp_uri = totp.provisioning_uri(name=request.user.email, issuer_name='Kwentas Klaras')

            # Save the secret to the user profile
            user_profile.totp_secret = secret
            user_profile.save()

            # Generate QR code
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(otp_uri)
            qr.make(fit=True)
            img = qr.make_image(fill='black', back_color='white')

            # Save QR code to a file-like object
            buffer = BytesIO()
            img.save(buffer, format='PNG')
            buffer.seek(0)
            user_profile.qr_code.save('qr_code.png', ContentFile(buffer.read()), save=True)
        
        user_profile.save()
        return JsonResponse({'success': True})

    return render(request, 'enable_2fa.html', {'qr_code': request.user.userprofile.qr_code.url})

def verify_otp(request):
    if request.method == 'POST':
        otp_code = request.POST.get('otp')
        user = request.user

        # Ensure 2FA is enabled
        if not user.userprofile.two_factor_enabled:
            messages.error(request, '2FA is not enabled for your account.')
            return redirect('homepage')

        # Check if TOTP secret exists
        if not user.userprofile.totp_secret:
            messages.error(request, 'No TOTP secret found.')
            return redirect('homepage')

        # Create TOTP object and verify OTP
        totp = pyotp.TOTP(user.userprofile.totp_secret)
        if totp.verify(otp_code):
            request.session['otp_verified'] = True
            request.session.pop('otp_required', None)  # Clear OTP requirement
            messages.success(request, 'OTP Verified Successfully!')
            return redirect('homepage')
        else:
            messages.error(request, 'Invalid OTP, please try again.')

    return render(request, 'KwentasApp/verify_otp.html')

@login_required
def get_2fa_status(request):
    user_profile = request.user.userprofile
    return JsonResponse({'is_2fa_enabled': user_profile.two_factor_enabled})

def validate_password(request):
    if request.method == 'POST':
        password = request.POST.get('password')
        user = request.user  # Assuming the user is already authenticated

        if user.is_authenticated:
            # Authenticate the user with the provided password
            user_check = authenticate(request, username=user.username, password=password)
            if user_check is not None:
                return JsonResponse({'valid': True})  # Password is correct
            else:
                return JsonResponse({'valid': False})  # Incorrect password
        else:
            return JsonResponse({'valid': False, 'error': 'User not authenticated'})
    return JsonResponse({'valid': False, 'error': 'Invalid request method'})


from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt


@require_POST
@csrf_exempt  # Exempting CSRF for testing (remove for production)
def delete_uploaded_file(request, file_id):
    try:
        # Use 'file_id' as the field name in the query
        file = UploadedFileData.objects.get(file_id=file_id)
        file.delete()
        return JsonResponse({'success': True})
    except UploadedFileData.DoesNotExist:
        return JsonResponse({'error': 'File not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)